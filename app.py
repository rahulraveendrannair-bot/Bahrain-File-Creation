"""
BH Sanctions List Updater — Web App (Offline / No API Key)
============================================================
Run locally:
    pip install flask openpyxl
    python app.py

Then open: http://localhost:5000
"""

import io
import json
import os
from datetime import datetime

from flask import Flask, jsonify, render_template_string, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Font

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB max upload

COLUMNS = [
    "NAME", "AKA", "FOREIGN_SCRIPT", "SEX", "DOB", "POB",
    "NATIONALITY", "OTHER_INFO", "ADD", "ADD_COUNTRY",
    "TITLE", "CITIZENSHIP", "REMARK",
]

# ── HTML Template ─────────────────────────────────────────────────────────
HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>BH Sanctions List Updater</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet"/>
<style>
:root{--bg:#0a0c0f;--s1:#111318;--s2:#181c23;--bd:#1e2330;--bd2:#252c3a;--ac:#00e5a0;--ac2:#0084ff;--tx:#e8eaf0;--t2:#7a8299;--t3:#4a5168;--err:#ff4d6a;--warn:#ffaa00}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--tx);font-family:'DM Mono',monospace;min-height:100vh}
body::before{content:'';position:fixed;inset:0;background-image:linear-gradient(rgba(0,229,160,.03) 1px,transparent 1px),linear-gradient(90deg,rgba(0,229,160,.03) 1px,transparent 1px);background-size:40px 40px;pointer-events:none;z-index:0}
.wrap{max-width:860px;margin:0 auto;padding:40px 24px 80px;position:relative;z-index:1}
header{display:flex;align-items:flex-start;gap:16px;margin-bottom:36px}
.logo{width:44px;height:44px;background:var(--ac);border-radius:8px;display:flex;align-items:center;justify-content:center;flex-shrink:0;margin-top:4px}
h1{font-family:'Syne',sans-serif;font-size:23px;font-weight:800;color:var(--tx);letter-spacing:-.5px;line-height:1.2}
h1 span{color:var(--ac)}
.sub{color:var(--t2);font-size:12px;margin-top:5px;line-height:1.6}
.badge-offline{background:rgba(0,132,255,.12);color:var(--ac2);border:1px solid rgba(0,132,255,.25);border-radius:4px;padding:2px 8px;font-size:10px;font-family:'Syne',sans-serif;font-weight:700;margin-left:8px;vertical-align:middle}
.grid3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px;margin-bottom:16px}
.dz{background:var(--s1);border:1.5px dashed var(--bd2);border-radius:10px;padding:22px 14px;text-align:center;cursor:pointer;transition:all .2s;position:relative}
.dz:hover{border-color:var(--ac);background:rgba(0,229,160,.06)}
.dz.ok{border-color:var(--ac);border-style:solid;background:rgba(0,229,160,.04)}
.dz input{position:absolute;inset:0;opacity:0;cursor:pointer;font-size:0}
.dico{width:36px;height:36px;margin:0 auto 9px;background:var(--s2);border-radius:7px;display:flex;align-items:center;justify-content:center;border:1px solid var(--bd2);font-size:18px}
.dlbl{font-family:'Syne',sans-serif;font-size:11px;font-weight:700;color:var(--tx);margin-bottom:2px}
.dsub{font-size:10px;color:var(--t3)}
.dfn{font-size:10px;color:var(--ac);margin-top:6px;display:none;word-break:break-all}
.dz.ok .dfn{display:block}
.dz.ok .dsub{display:none}
.tick{position:absolute;top:7px;right:9px;color:var(--ac);font-size:12px;display:none;font-weight:bold}
.dz.ok .tick{display:block}
.arow{display:flex;gap:10px;margin-bottom:22px;align-items:center;flex-wrap:wrap}
.btn{background:var(--ac);color:var(--bg);border:none;border-radius:10px;padding:12px 28px;font-family:'Syne',sans-serif;font-size:13px;font-weight:700;cursor:pointer;display:flex;align-items:center;gap:8px;transition:opacity .2s}
.btn:hover:not(:disabled){opacity:.85}
.btn:disabled{opacity:.3;cursor:not-allowed}
.btn2{background:var(--ac2);color:#fff;border:none;border-radius:10px;padding:12px 22px;font-family:'Syne',sans-serif;font-size:13px;font-weight:700;cursor:pointer;display:flex;align-items:center;gap:8px;text-decoration:none}
.btn2:hover{opacity:.88}
.btn3{background:transparent;color:var(--t2);border:1.5px solid var(--bd2);border-radius:10px;padding:11px 18px;font-family:'Syne',sans-serif;font-size:12px;font-weight:600;cursor:pointer}
.btn3:hover{border-color:var(--t2);color:var(--tx)}
.spin{width:14px;height:14px;border:2px solid rgba(10,12,15,.2);border-top-color:var(--bg);border-radius:50%;animation:sp .7s linear infinite;display:none;flex-shrink:0}
@keyframes sp{to{transform:rotate(360deg)}}
.term{background:var(--s1);border:1px solid var(--bd);border-radius:10px;padding:16px 18px;max-height:200px;overflow-y:auto;margin-bottom:16px;display:none}
.ll{display:flex;gap:8px;font-size:11px;line-height:1.9}
.lt{color:var(--t3);flex-shrink:0}
.li{color:var(--t2)}.ls{color:var(--ac)}.le{color:var(--err)}.lw{color:var(--warn)}.lx{color:var(--ac2);font-weight:bold}
.ban{border-radius:10px;padding:13px 18px;display:none;align-items:center;gap:12px;margin-bottom:16px}
.bok{background:rgba(0,229,160,.08);border:1px solid rgba(0,229,160,.22)}
.ber{background:rgba(255,77,106,.08);border:1px solid rgba(255,77,106,.22)}
.bt{font-family:'Syne',sans-serif;font-size:13px;font-weight:700}
.bd{font-size:11px;color:var(--t2);margin-top:2px}
.info-box{background:var(--s1);border:1px solid var(--bd2);border-radius:10px;padding:14px 18px;margin-bottom:22px;font-size:11px;color:var(--t2);line-height:1.8}
.info-box strong{color:var(--tx);font-weight:500}
.info-box code{background:var(--s2);border:1px solid var(--bd2);border-radius:4px;padding:1px 6px;font-size:10px;color:var(--ac)}
#dlrow{display:none;gap:10px;align-items:center;flex-wrap:wrap}
::-webkit-scrollbar{width:5px}
::-webkit-scrollbar-track{background:var(--s1)}
::-webkit-scrollbar-thumb{background:var(--bd2);border-radius:3px}
</style>
</head>
<body>
<div class="wrap">

<header>
  <div class="logo">
    <svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="#0a0c0f" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="12" y1="18" x2="12" y2="12"/><line x1="9" y1="15" x2="15" y2="15"/></svg>
  </div>
  <div>
    <h1>BH Sanctions <span>List Updater</span> <span class="badge-offline">OFFLINE</span></h1>
    <p class="sub">Upload the Gazette PDF + BH-TL-INDIVIDUALS.xlsx + extracted individuals JSON.<br>The app appends new records to the XLSX and lets you download the updated file.</p>
  </div>
</header>

<div class="info-box">
  <strong>How to prepare the JSON file:</strong><br>
  Extract individuals from the Gazette PDF using Claude AI in this chat, then save the result as a <code>.json</code> file.<br>
  The JSON must be an array of objects with keys: <code>NAME AKA FOREIGN_SCRIPT SEX DOB POB NATIONALITY OTHER_INFO ADD ADD_COUNTRY TITLE CITIZENSHIP REMARK</code>
</div>

<div class="grid3">
  <div class="dz" id="pdfZone">
    <input type="file" id="pdfInput" accept=".pdf">
    <span class="tick">✓</span>
    <div class="dico">📄</div>
    <div class="dlbl">Official Gazette PDF</div>
    <div class="dsub">Click or drag .pdf</div>
    <div class="dfn" id="pdfName"></div>
  </div>
  <div class="dz" id="xlsxZone">
    <input type="file" id="xlsxInput" accept=".xlsx,.xls">
    <span class="tick">✓</span>
    <div class="dico">📊</div>
    <div class="dlbl">BH-TL-INDIVIDUALS.xlsx</div>
    <div class="dsub">Click or drag .xlsx</div>
    <div class="dfn" id="xlsxName"></div>
  </div>
  <div class="dz" id="jsonZone">
    <input type="file" id="jsonInput" accept=".json">
    <span class="tick">✓</span>
    <div class="dico">{ }</div>
    <div class="dlbl">Individuals JSON</div>
    <div class="dsub">Click or drag .json</div>
    <div class="dfn" id="jsonName"></div>
  </div>
</div>

<div class="arow">
  <button class="btn" id="runBtn" disabled onclick="run()">
    <span class="spin" id="sp"></span>
    <svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round" id="playico"><polygon points="5 3 19 12 5 21 5 3"/></svg>
    <span id="btxt">Update XLSX</span>
  </button>
</div>

<div class="term" id="term"></div>

<div class="ban" id="ban">
  <span id="bi" style="font-size:20px"></span>
  <div><div class="bt" id="btt"></div><div class="bd" id="bds"></div></div>
</div>

<div id="dlrow">
  <a class="btn2" id="dlLink" href="#" download>
    <svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
    Download Updated XLSX
  </a>
  <button class="btn3" onclick="rst()">Process Another File</button>
</div>

</div>
<script>
let files = {pdf: null, xlsx: null, json: null};

function setupDrop(zoneId, inputId, key, accept) {
  const zone = document.getElementById(zoneId);
  const inp  = document.getElementById(inputId);
  zone.addEventListener('dragover', e => { e.preventDefault(); zone.classList.add('drag'); });
  zone.addEventListener('dragleave', () => zone.classList.remove('drag'));
  zone.addEventListener('drop', e => { e.preventDefault(); zone.classList.remove('drag'); if (e.dataTransfer.files[0]) setFile(e.dataTransfer.files[0], key, zoneId, inputId); });
  inp.addEventListener('change', () => { if (inp.files[0]) setFile(inp.files[0], key, zoneId, inputId); });
}
function setFile(f, key, zoneId, nameId) {
  files[key] = f;
  document.getElementById(zoneId).classList.add('ok');
  document.getElementById(nameId.replace('Zone','Name')).textContent = f.name;
  checkReady();
}
function checkReady() {
  document.getElementById('runBtn').disabled = !(files.pdf && files.xlsx && files.json);
}
setupDrop('pdfZone',  'pdfInput',  'pdf',  '.pdf');
setupDrop('xlsxZone', 'xlsxInput', 'xlsx', '.xlsx,.xls');
setupDrop('jsonZone', 'jsonInput', 'json', '.json');

function ts() { return new Date().toTimeString().slice(0,8); }
function lg(m, c) {
  const t = document.getElementById('term'); t.style.display = 'block';
  const d = document.createElement('div'); d.className = 'll';
  d.innerHTML = `<span class="lt">[${ts()}]</span><span class="${c||'li'}">${m}</span>`;
  t.appendChild(d); t.scrollTop = t.scrollHeight;
}
function setBusy(msg) {
  document.getElementById('sp').style.display = 'inline-block';
  document.getElementById('playico').style.display = 'none';
  document.getElementById('btxt').textContent = msg;
  document.getElementById('runBtn').disabled = true;
}
function setIdle() {
  document.getElementById('sp').style.display = 'none';
  document.getElementById('playico').style.display = 'inline';
  document.getElementById('btxt').textContent = 'Update XLSX';
}
function showBan(ok, title, desc) {
  const b = document.getElementById('ban');
  b.className = 'ban ' + (ok ? 'bok' : 'ber'); b.style.display = 'flex';
  document.getElementById('bi').textContent = ok ? '✅' : '❌';
  document.getElementById('btt').style.color = ok ? '#00e5a0' : '#ff4d6a';
  document.getElementById('btt').textContent = title;
  document.getElementById('bds').textContent = desc;
}

async function run() {
  document.getElementById('term').innerHTML = '';
  document.getElementById('term').style.display = 'none';
  document.getElementById('ban').style.display = 'none';
  document.getElementById('dlrow').style.display = 'none';

  setBusy('Uploading files...');
  lg('Preparing files...', 'li');

  const fd = new FormData();
  fd.append('pdf',  files.pdf);
  fd.append('xlsx', files.xlsx);
  fd.append('json', files.json);

  lg('Sending to server...', 'lx');

  try {
    const res = await fetch('/update', { method: 'POST', body: fd });
    const data = await res.json();

    if (!res.ok || data.error) {
      lg('Error: ' + (data.error || 'Unknown error'), 'le');
      showBan(false, 'Failed', data.error || 'Server error');
      setIdle(); return;
    }

    data.log.forEach(entry => lg(entry.msg, entry.cls));

    showBan(true, 'Updated Successfully',
      `${data.added} individual(s) added. ${data.skipped ? data.skipped + ' duplicate(s) skipped.' : ''}`);

    const dlLink = document.getElementById('dlLink');
    dlLink.href = '/download/' + data.filename;
    dlLink.download = data.filename;
    document.getElementById('dlrow').style.display = 'flex';

  } catch (e) {
    lg('Request failed: ' + e.message, 'le');
    showBan(false, 'Request Failed', e.message);
  }

  setIdle();
}

function rst() {
  files = {pdf: null, xlsx: null, json: null};
  ['pdfZone','xlsxZone','jsonZone'].forEach(id => document.getElementById(id).classList.remove('ok'));
  ['pdfName','xlsxName','jsonName'].forEach(id => document.getElementById(id).textContent = '');
  ['pdfInput','xlsxInput','jsonInput'].forEach(id => document.getElementById(id).value = '');
  document.getElementById('runBtn').disabled = true;
  document.getElementById('term').innerHTML = '';
  document.getElementById('term').style.display = 'none';
  document.getElementById('ban').style.display = 'none';
  document.getElementById('dlrow').style.display = 'none';
  setIdle();
}
</script>
</body>
</html>"""


# ── In-memory storage for generated files ─────────────────────────────────
_output_store: dict[str, bytes] = {}


@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/update", methods=["POST"])
def update():
    logs = []

    def lg(msg, cls="li"):
        logs.append({"msg": msg, "cls": cls})

    # ── Validate uploads ───────────────────────────────────────────────────
    for field in ("pdf", "xlsx", "json"):
        if field not in request.files or request.files[field].filename == "":
            return jsonify({"error": f"Missing file: {field}"}), 400

    xlsx_file = request.files["xlsx"]
    json_file = request.files["json"]

    # ── Parse JSON ─────────────────────────────────────────────────────────
    try:
        raw = json_file.read().decode("utf-8")
        individuals = json.loads(raw)
        if not isinstance(individuals, list):
            raise ValueError("JSON must be an array")
        individuals = [
            {k: (row.get(k, "") or "") for k in COLUMNS}
            for row in individuals
            if isinstance(row, dict)
        ]
        lg(f"JSON loaded: {len(individuals)} individual(s)", "ls")
    except Exception as e:
        return jsonify({"error": f"Invalid JSON: {e}"}), 400

    if not individuals:
        return jsonify({"error": "No valid individuals found in JSON"}), 400

    # ── Load XLSX ──────────────────────────────────────────────────────────
    try:
        wb = load_workbook(xlsx_file)
        ws = wb.active
        existing = {
            str(row[0]).strip().lower()
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row and row[0]
        }
        lg(f"XLSX loaded: {ws.max_row - 1} existing records", "ls")
    except Exception as e:
        return jsonify({"error": f"Cannot read XLSX: {e}"}), 400

    # ── Append records ─────────────────────────────────────────────────────
    added = skipped = 0
    for person in individuals:
        name = person.get("NAME", "").strip()
        key = name.lower()
        if not key or key in existing:
            lg(f"Skipped duplicate: {name}", "lw")
            skipped += 1
            continue
        row_num = ws.max_row + 1
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            val = person.get(col_name) or None
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=11)
        existing.add(key)
        lg(f"Added: {name}", "ls")
        added += 1

    # Preserve column widths
    for col_letter, width in {"A":49.36,"B":39.82,"C":14.82,"E":14.18,"F":11.82,"G":8.54,"H":33.45,"K":7.18,"L":6.82,"M":13.18}.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"

    # ── Save to memory ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    base = os.path.splitext(xlsx_file.filename)[0]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_filename = f"{base}-UPDATED-{ts}.xlsx"
    _output_store[out_filename] = buf.read()

    lg(f"Done — {added} added, {skipped} skipped", "lx")

    return jsonify({
        "added": added,
        "skipped": skipped,
        "filename": out_filename,
        "log": logs,
    })


@app.route("/download/<filename>")
def download(filename):
    if filename not in _output_store:
        return "File not found", 404
    return send_file(
        io.BytesIO(_output_store[filename]),
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    print("\n" + "=" * 50)
    print("  BH Sanctions List Updater")
    print("  Running at: http://localhost:5000")
    print("=" * 50 + "\n")
    app.run(debug=True, port=5000)
